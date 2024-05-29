#!/usr/bin/env python3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os

def process_procedures(input_file):

    base_name = os.path.splitext(os.path.basename(input_file))[0]

    data_frame = pd.read_excel(input_file, sheet_name='Master')
    data_frame['DATE'] = pd.to_datetime(data_frame['DATE']).dt.date
    data_frame['COST_EXT'] = data_frame['COST_EXT'].replace(r'[\$,]', '', regex=True).astype(float)

    aggregation_dictionary = {
        'COST_EXT': 'sum',
        'PRI_SURG_NAME': 'first',
        'DATE': lambda x: 'TOTAL'
    }

    total_cost_ext = data_frame.groupby('LOG_ID').agg(aggregation_dictionary).reset_index()

    data_frame = pd.concat([data_frame, total_cost_ext], ignore_index=True)
    data_frame = data_frame.sort_values(by=['LOG_ID', 'DATE']).reset_index(drop=True)
    data_frame['COST_EXT'] = data_frame['COST_EXT'].apply(lambda x: f"${x:,.2f}")

    wb = Workbook()
    ws = wb.active

    current_log_id = None
    current_date = None
    current_physicians_name = None
    log_id_start_row = None
    date_start_row = None
    physician_start_row = None

    for row_index, row in enumerate(dataframe_to_rows(data_frame, index=False, header=True), 1):
        for column_index, value in enumerate(row, 1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            if isinstance(value, str):
                cell.alignment = Alignment(wrapText=True)

            if column_index == 1:
                if value != current_log_id:
                    if current_log_id is not None and log_id_start_row is not None and row_index > log_id_start_row:
                        ws.merge_cells(start_row=log_id_start_row, start_column=1, end_row=row_index-1, end_column=1)
                    current_log_id = value
                    log_id_start_row = row_index

            elif column_index == 2:
                if value != current_date:
                    if current_date is not None and date_start_row is not None and row_index > date_start_row:
                        ws.merge_cells(start_row=date_start_row, start_column=2, end_row=row_index-1, end_column=2)
                    current_date = value
                    date_start_row = row_index

            elif column_index == 3:
                if value != current_physicians_name:
                    if current_physicians_name is not None and physician_start_row is not None and row_index > physician_start_row:
                        ws.merge_cells(start_row=physician_start_row, start_column=3, end_row=row_index-1, end_column=3)
                    current_physicians_name = value
                    physician_start_row = row_index

    if current_log_id is not None and log_id_start_row is not None and row_index >= log_id_start_row:
        ws.merge_cells(start_row=log_id_start_row, start_column=1, end_row=row_index, end_column=1)
    if current_date is not None and date_start_row is not None and row_index >= date_start_row:
        ws.merge_cells(start_row=date_start_row, start_column=2, end_row=row_index, end_column=2)
    if current_physicians_name is not None and physician_start_row is not None and row_index >= physician_start_row:
        ws.merge_cells(start_row=physician_start_row, start_column=3, end_row=row_index, end_column=3)


    output_file = f"{base_name}_processed.xlsx"
    wb.save(output_file)

    print(f"Output saved to '{output_file}'")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: totalproc <input_file>")
        sys.exit(1)

    input_file = sys.argv[1]
    process_procedures(input_file)
